import xml.etree.ElementTree as ET
import openpyxl

# Define the input and output files
input_file = 'go_obo.xml'
output_file = 'autophagosome.xlsx'

# Create a new workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Add headers to the worksheet
worksheet.cell(row=1, column=1, value='GO id')
worksheet.cell(row=1, column=2, value='Term name')
worksheet.cell(row=1, column=3, value='Definition string')
worksheet.cell(row=1, column=4, value='Number of child nodes')

# Parse the XML file
tree = ET.parse(input_file)
root = tree.getroot()

# Define a function to perform DFS on DAG and count child nodes
def dfs_count(node, visited):
    # Mark node as visited
    visited.add(node)
    # Count child nodes
    child_node_count = 0
    for neighbor in adj_list[node]:
        if neighbor not in visited:
            child_node_count += 1
            child_node_count += dfs_count(neighbor, visited)
    return child_node_count

# Create adjacency list for DAG
adj_list = {}
for term in root.findall('term'):
    term_id = term.find('id').text
    adj_list[term_id] = []
    for is_a in term.findall('is_a'):
        parent_id = is_a.text
        adj_list[term_id].append(parent_id)

# Loop through all the <term> elements
row = 2
for term in root.findall('term'):
    # Check if the <defstr> element contains the word 'autophagosome'
    if 'autophagosome' in term.find('def').find('defstr').text:
        # Extract the required information
        go_id = term.find('id').text
        term_name = term.find('name').text
        def_str = term.find('def').find('defstr').text
        child_nodes = dfs_count(go_id, set()) - 1

        # Add the information to the worksheet
        worksheet.cell(row=row, column=1, value=go_id)
        worksheet.cell(row=row, column=2, value=term_name)
        worksheet.cell(row=row, column=3, value=def_str)
        worksheet.cell(row=row, column=4, value=child_nodes)

        # Increment the row counter
        row += 1

# Save the workbook
workbook.save(output_file)